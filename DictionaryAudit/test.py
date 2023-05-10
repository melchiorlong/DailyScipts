import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple


# 1.打开 Excel 表格并获取表格名称
workbook = load_workbook(filename="/Users/tianlong/Desktop/中信JD.xlsx")
sheet_names_list = workbook.sheetnames

# 2.通过 sheet 名称获取表格
for sheet_name in sheet_names_list:
	sheet = workbook[sheet_name]
	print(sheet)
	# 3.获取表格的尺寸大小(几行几列数据) 这里所说的尺寸大小，指的是 excel 表格中的数据有几行几列，针对的是不同的 sheet 而言。
	print(sheet.dimensions)
	# 4.获取表格内某个格子的数据
	# 1 sheet["A1"]方式
	cell1 = sheet["A1"]
	cell2 = sheet["C6"]
	print(cell1.value, cell2.value)
	"""
	workbook.active 打开激活的表格; sheet["A1"] 获取 A1 格子的数据; cell.value 获取格子中的值;
	"""
	# 4.2sheet.cell(row=, column=)方式
	cell1 = sheet.cell(row=1, column=1)
	cell2 = sheet.cell(row=11, column=3)
	print(cell1.value, cell2.value)

	# 5. 获取一系列格子
	# 获取 A1:C2 区域的值
	cell = sheet["A1:C2"]
	print(cell)
	for i in cell:
		for j in i:
			print(j.value)
			print(j.coordinate)
			print(coordinate_to_tuple(j.coordinate))
	print("-----")


